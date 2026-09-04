"""
Massa-import vanuit een Excel-bestand (.xlsx).

Bewust met openpyxl in read-only modus en zonder pandas: pandas is een zware
afhankelijkheid die op een Raspberry Pi traag installeert en veel geheugen
vraagt, terwijl we hier enkel rijen willen doorlopen (vereiste 3c).
"""
from openpyxl import load_workbook

# Kolomnaam in het Excel-bestand -> veldnaam op het Media-model.
COLUMN_MAP = {
    "type": "media_type_code", "mediatype": "media_type_code",
    "titel": "title", "title": "title",
    "reeks": "series", "series": "series",
    "nummer in de reeks": "series_number", "reeksnummer": "series_number", "series_number": "series_number",
    "auteur": "author", "auteur / tekenaar": "author", "tekenaar": "author", "author": "author",
    "collectie": "collection", "collection": "collection",
    "nummer in de collectie": "collection_number",
    "nummer van de druk": "print_number", "druk": "print_number",
    "eigenaar": "owner_name", "owner": "owner_name",
    "dubbel": "is_duplicate",
    "hardcover": "is_hardcover",
    "staat": "condition", "conditie": "condition",
    "commentaar": "comment", "comment": "comment", "opmerking": "comment",
    "muzikant": "musician", "musician": "musician", "artiest": "musician",
    "jaar": "year", "year": "year",
    "taal audio": "audio_language", "audio": "audio_language",
    "taal ondertiteling": "subtitle_language", "ondertiteling": "subtitle_language",
    "barcode": "barcode", "isbn": "barcode", "ean": "barcode",
    "waarde": "estimated_value", "geschatte waarde": "estimated_value",
}

TRUE_VALUES = {"1", "true", "ja", "yes", "x", "waar"}
MAX_ROWS = 20000  # veiligheidsgrens tegen een gigantisch bestand


def read_import_file(filepath):
    """
    Leest het bestand rij per rij in en normaliseert de kolomnamen
    (hoofdletterongevoelig). Geeft een lijst dicts terug, klaar om als
    Media-records aangemaakt te worden. Onbekende kolommen worden genegeerd.
    """
    workbook = load_workbook(filepath, read_only=True, data_only=True)
    sheet = workbook.active

    rows = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        workbook.close()
        return []

    headers = [str(cell).strip().lower() if cell is not None else "" for cell in header_row]

    records = []
    for index, row in enumerate(rows):
        if index >= MAX_ROWS:
            break
        record = {}
        for header, value in zip(headers, row):
            field = COLUMN_MAP.get(header)
            if not field or value is None or str(value).strip() == "":
                continue
            if field in ("is_duplicate", "is_hardcover"):
                value = str(value).strip().lower() in TRUE_VALUES
            record[field] = value
        if record.get("title") or record.get("media_type_code"):
            records.append(record)

    workbook.close()
    return records
