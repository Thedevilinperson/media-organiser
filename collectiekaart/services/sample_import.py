"""
Bouwt het voorbeeldbestand voor de massa-import.

Het bestand wordt bij het downloaden aangemaakt uit dezelfde kolomtabel als de
importer zelf gebruikt. Zo kan het nooit uit de pas lopen met wat de app
werkelijk inleest.
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

KOLOMMEN = [
    "type", "titel", "reeks", "nummer in de reeks", "auteur", "collectie",
    "nummer in de collectie", "nummer van de druk", "eigenaar", "dubbel",
    "hardcover", "staat", "commentaar", "muzikant", "jaar", "taal audio",
    "taal ondertiteling", "barcode", "waarde",
]

VOORBEELDRIJEN = [
    ["strip", "De Zwarte Zon", "Thorgal", 12, "Rosinski / Van Hamme", "Integraal",
     3, 2, "Jan", "nee", "ja", "goede staat", "kaft licht verkleurd", "", 1997,
     "", "", "9789034306715", 14.50],
    ["strip", "Het Eiland", "Suske en Wiske", 301, "Willy Vandersteen", "",
     "", 1, "Jan", "nee", "nee", "nieuwstaat", "", "", 2008, "", "", "", 6.00],
    ["boek", "De Ontdekking van de Hemel", "", "", "Harry Mulisch", "",
     "", 4, "Marie", "nee", "ja", "goede staat", "", "", 1992, "", "",
     "9789023456789", 12.00],
    ["cd", "Kind of Blue", "", "", "", "", "", "", "Jan", "nee", "", "", "",
     "Miles Davis", 1959, "", "", "", 18.00],
    ["dvd", "Amélie", "", "", "", "", "", "", "Marie", "nee", "", "", "", "",
     2001, "Frans", "Nederlands, Engels", "", 8.50],
]

TOELICHTING = [
    ("Zo gebruik je dit bestand", None),
    ("De eerste rij met kolomnamen moet blijven staan. Kolomnamen zijn "
     "hoofdletterongevoelig; kolommen die de app niet kent, worden genegeerd.", None),
    ("De kolom 'type' bevat de interne code van een bestaand mediatype. "
     "Die codes vind je in Instellingen, bijvoorbeeld strip, boek, cd of dvd. "
     "Rijen met een onbekend type worden overgeslagen en na de import opgesomd.", None),
    ("Voor 'dubbel' en 'hardcover' mag je ja, nee, 1, 0, waar of x gebruiken.", None),
    ("Voor 'staat' gebruik je een van deze waarden: slechte staat, redelijke staat, "
     "goede staat, bijna nieuwstaat, nieuwstaat.", None),
    ("Eigenaars die nog niet bestaan, maakt de app tijdens de import zelf aan.", None),
    ("Verwijder gerust de voorbeeldrijen en zet er je eigen collectie in.", None),
]


def build_sample_workbook():
    """Geeft het voorbeeldbestand terug als BytesIO, klaar om te versturen."""
    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Collectie"
    sheet.append(KOLOMMEN)

    kop = Font(bold=True, color="FFFFFF")
    vulling = PatternFill("solid", fgColor="2C5A56")
    for cel in sheet[1]:
        cel.font = kop
        cel.fill = vulling
        cel.alignment = Alignment(vertical="center")

    for rij in VOORBEELDRIJEN:
        sheet.append(rij)

    for index, naam in enumerate(KOLOMMEN, start=1):
        letter = sheet.cell(row=1, column=index).column_letter
        sheet.column_dimensions[letter].width = max(14, len(naam) + 3)
    sheet.freeze_panes = "A2"

    uitleg = workbook.create_sheet("Toelichting")
    uitleg.column_dimensions["A"].width = 110
    for rij, (tekst, _) in enumerate(TOELICHTING, start=1):
        cel = uitleg.cell(row=rij, column=1, value=tekst)
        cel.alignment = Alignment(wrap_text=True, vertical="top")
        if rij == 1:
            cel.font = Font(bold=True, size=13)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
